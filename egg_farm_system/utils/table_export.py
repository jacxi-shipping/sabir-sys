"""
Data export utilities for tables.

Provides easy export of table data to CSV and Excel formats.
"""
import csv
from pathlib import Path
from typing import List, Optional
from datetime import datetime


class DataExporter:
    """Export table data to CSV or Excel.
    
    Usage:
        exporter = DataExporter()
        exporter.export_to_csv(headers, rows, "output.csv")
    """
    
    @staticmethod
    def export_to_csv(headers: List[str], rows: List[List[str]], filepath: str) -> bool:
        """Export data to CSV file.
        
        Args:
            headers: List of column headers
            rows: List of data rows
            filepath: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                # Write headers
                writer.writerow(headers)
                # Write data rows
                writer.writerows(rows)
            return True
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    @staticmethod
    def export_to_excel(headers: List[str], rows: List[List[str]], filepath: str, sheet_name: str = "Data") -> bool:
        """Export data to Excel file.
        
        Args:
            headers: List of column headers
            rows: List of data rows
            filepath: Output file path
            sheet_name: Name of the worksheet
            
        Returns:
            True if successful, False otherwise
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers with formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    # Remove HTML tags if present (like <b> tags)
                    clean_value = str(value).replace('<b>', '').replace('</b>', '')
                    ws.cell(row=row_idx, column=col_idx, value=clean_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            return True
        except ImportError:
            print("openpyxl not installed. Cannot export to Excel.")
            return False
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    @staticmethod
    def get_default_filename(prefix: str, format: str = "csv") -> str:
        """Generate default filename with timestamp.
        
        Args:
            prefix: Prefix for filename (e.g., "sales", "parties")
            format: File format ("csv" or "xlsx")
            
        Returns:
            Filename with timestamp
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{format}"


class TableExportMixin:
    """Mixin to add export functionality to table widgets.
    
    Usage:
        class MyTableWidget(QWidget, TableExportMixin):
            def __init__(self):
                super().__init__()
                self.add_export_buttons()
    """
    
    def add_export_buttons(self, layout, table_widget, data_name="data"):
        """Add export buttons to a layout.
        
        Args:
            layout: QLayout to add buttons to
            table_widget: DataTableWidget instance
            data_name: Name for exported file (e.g., "sales", "parties")
        """
        from PySide6.QtWidgets import QPushButton, QFileDialog, QMessageBox
        
        # CSV export button
        csv_btn = QPushButton("📊 Export to CSV")
        csv_btn.setToolTip("Export table data to CSV file")
        csv_btn.clicked.connect(lambda: self._export_table_csv(table_widget, data_name))
        layout.addWidget(csv_btn)
        
        # Excel export button
        excel_btn = QPushButton("📊 Export to Excel")
        excel_btn.setToolTip("Export table data to Excel file")
        excel_btn.clicked.connect(lambda: self._export_table_excel(table_widget, data_name))
        layout.addWidget(excel_btn)
    
    def _export_table_csv(self, table_widget, data_name):
        """Export table to CSV"""
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        
        # Get headers and rows from table
        headers = self._get_table_headers(table_widget)
        rows = self._get_table_rows(table_widget)
        
        if not rows:
            QMessageBox.information(self, "No Data", "No data to export")
            return
        
        # Ask user for file location
        default_name = DataExporter.get_default_filename(data_name, "csv")
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            default_name,
            "CSV Files (*.csv)"
        )
        
        if filepath:
            if DataExporter.export_to_csv(headers, rows, filepath):
                QMessageBox.information(self, "Success", f"Data exported to {filepath}")
            else:
                QMessageBox.critical(self, "Error", "Failed to export data")
    
    def _export_table_excel(self, table_widget, data_name):
        """Export table to Excel"""
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        
        # Get headers and rows from table
        headers = self._get_table_headers(table_widget)
        rows = self._get_table_rows(table_widget)
        
        if not rows:
            QMessageBox.information(self, "No Data", "No data to export")
            return
        
        # Ask user for file location
        default_name = DataExporter.get_default_filename(data_name, "xlsx")
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            default_name,
            "Excel Files (*.xlsx)"
        )
        
        if filepath:
            if DataExporter.export_to_excel(headers, rows, filepath):
                QMessageBox.information(self, "Success", f"Data exported to {filepath}")
            else:
                QMessageBox.critical(self, "Error", "Failed to export data")
    
    def _get_table_headers(self, table_widget):
        """Extract headers from DataTableWidget"""
        # Assuming table_widget has a method to get headers
        if hasattr(table_widget, 'headers'):
            return table_widget.headers
        # Fallback: extract from table
        return []
    
    def _get_table_rows(self, table_widget):
        """Extract rows from DataTableWidget"""
        # This needs to be customized based on how the table stores data
        # For now, return empty list - subclasses should override
        return []
