"""
Template generator for bulk data import
"""
import csv
from pathlib import Path
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available, Excel templates disabled")


class TemplateGenerator:
    """Generate Excel/CSV templates for bulk import"""
    
    TEMPLATES = {
        'parties': {
            'columns': ['name', 'phone', 'address', 'notes'],
            'required': ['name'],
            'example': ['ABC Poultry Traders', '+93701234567', '123 Main St, Kabul', 'Premium customer']
        },
        'raw_materials': {
            'columns': ['name', 'unit', 'low_stock_alert', 'supplier_name', 'notes'],
            'required': ['name', 'unit'],
            'example': ['Corn', 'kg', '1000', 'ABC Suppliers', 'Yellow corn grade A']
        },
        'expenses': {
            'columns': ['date', 'farm_name', 'category', 'amount_afg', 'amount_usd', 'description', 'payment_method'],
            'required': ['date', 'farm_name', 'category', 'amount_afg'],
            'example': ['2026-01-15', 'Farm 1', 'Labor', '5000', '64', 'Daily wages', 'Cash']
        },
        'employees': {
            'columns': ['full_name', 'job_title', 'hire_date', 'salary_amount', 'salary_period', 'is_active'],
            'required': ['full_name', 'job_title', 'salary_amount', 'salary_period'],
            'example': ['Ahmad Khan', 'Farm Manager', '2026-01-01', '15000', 'Monthly', 'Yes']
        },
    }
    
    @staticmethod
    def get_available_templates():
        """Get list of available template types"""
        return list(TemplateGenerator.TEMPLATES.keys())
    
    @staticmethod
    def generate_csv_template(entity_type: str, filepath: str) -> bool:
        """
        Generate CSV template with headers
        
        Args:
            entity_type: Type of entity (parties, raw_materials, etc.)
            filepath: Path where template should be saved
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            if entity_type not in TemplateGenerator.TEMPLATES:
                logger.error(f"Unknown template type: {entity_type}")
                return False
            
            template = TemplateGenerator.TEMPLATES[entity_type]
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write header row
                writer.writerow(template['columns'])
                
                # Write example row
                writer.writerow(template['example'])
            
            logger.info(f"CSV template generated: {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating CSV template: {e}")
            return False
    
    @staticmethod
    def generate_excel_template(entity_type: str, filepath: str) -> bool:
        """
        Generate Excel template with headers and formatting
        
        Args:
            entity_type: Type of entity (parties, raw_materials, etc.)
            filepath: Path where template should be saved
            
        Returns:
            bool: True if successful, False otherwise
        """
        if not EXCEL_AVAILABLE:
            logger.error("Excel templates not available - openpyxl not installed")
            return False
        
        try:
            if entity_type not in TemplateGenerator.TEMPLATES:
                logger.error(f"Unknown template type: {entity_type}")
                return False
            
            template = TemplateGenerator.TEMPLATES[entity_type]
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = entity_type.replace('_', ' ').title()
            
            # Header style
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Write headers
            for col_idx, column in enumerate(template['columns'], start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = column
                cell.fill = header_fill
                cell.font = header_font
                
                # Mark required fields
                if column in template['required']:
                    cell.value = f"{column} *"
            
            # Write example row
            for col_idx, value in enumerate(template['example'], start=1):
                ws.cell(row=2, column=col_idx, value=value)
            
            # Add instructions sheet
            ws_instructions = wb.create_sheet("Instructions")
            instructions = [
                ["Import Template Instructions"],
                [""],
                [f"Template Type: {entity_type.replace('_', ' ').title()}"],
                [""],
                ["Required Fields (marked with *):"],
            ]
            
            for field in template['required']:
                instructions.append([f"  - {field}"])
            
            instructions.extend([
                [""],
                ["Instructions:"],
                ["1. Fill in your data starting from row 2"],
                ["2. Do not modify the header row"],
                ["3. Required fields must not be empty"],
                ["4. Delete the example row before importing"],
                ["5. Save and use File > Import to upload"],
                [""],
                [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            ])
            
            for row_idx, row in enumerate(instructions, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws_instructions.cell(row=row_idx, column=col_idx, value=value)
            
            # Adjust column widths
            for ws_sheet in [ws, ws_instructions]:
                for column in ws_sheet.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"Excel template generated: {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating Excel template: {e}")
            return False
    
    @staticmethod
    def get_template_info(entity_type: str) -> dict:
        """Get information about a template"""
        if entity_type not in TemplateGenerator.TEMPLATES:
            return None
        
        template = TemplateGenerator.TEMPLATES[entity_type]
        return {
            'entity_type': entity_type,
            'columns': template['columns'],
            'required_fields': template['required'],
            'total_columns': len(template['columns']),
            'example_row': template['example']
        }
