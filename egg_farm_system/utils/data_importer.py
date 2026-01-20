"""
Data importer for bulk imports
"""
import csv
from datetime import datetime
from pathlib import Path
import logging

from egg_farm_system.database.db import DatabaseManager
from egg_farm_system.database.models import Party, RawMaterial, Expense, Employee, Farm, SalaryPeriod
from egg_farm_system.utils.data_validator import DataValidator

logger = logging.getLogger(__name__)

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available, Excel import disabled")


class ImportHistory:
    """Track import history"""
    
    def __init__(self):
        self.imports = []
    
    def add_import(self, user_id, entity_type, filepath, imported_count, error_count, imported_ids):
        """Add import to history"""
        self.imports.append({
            'timestamp': datetime.now(),
            'user_id': user_id,
            'entity_type': entity_type,
            'filepath': filepath,
            'imported_count': imported_count,
            'error_count': error_count,
            'imported_ids': imported_ids
        })
    
    def get_recent_imports(self, limit=10):
        """Get recent imports"""
        return sorted(self.imports, key=lambda x: x['timestamp'], reverse=True)[:limit]


class DataImporter:
    """Import data from Excel/CSV files"""
    
    def __init__(self, session=None):
        self.session = session or DatabaseManager.get_session()
        self.import_history = ImportHistory()
        self._owned_session = session is None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._owned_session and self.session:
            self.session.close()
    
    def _read_csv_file(self, filepath: str) -> list:
        """Read CSV file and return list of dictionaries"""
        data = []
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Skip empty rows
                    if any(row.values()):
                        data.append(row)
            return data
        except Exception as e:
            logger.error(f"Error reading CSV file: {e}")
            raise
    
    def _read_excel_file(self, filepath: str) -> list:
        """Read Excel file and return list of dictionaries"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not available")
        
        data = []
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    # Remove asterisk from required field markers
                    header = str(cell.value).replace(' *', '')
                    headers.append(header)
            
            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if any(row):
                    row_dict = {}
                    for idx, value in enumerate(row):
                        if idx < len(headers):
                            row_dict[headers[idx]] = value
                    data.append(row_dict)
            
            return data
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
    
    def _read_file(self, filepath: str) -> list:
        """Read file and return data based on extension"""
        file_path = Path(filepath)
        
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            return self._read_excel_file(filepath)
        elif file_path.suffix.lower() == '.csv':
            return self._read_csv_file(filepath)
        else:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")
    
    def import_parties(self, filepath: str, user_id: int = 1) -> dict:
        """Import parties from file"""
        try:
            # Read file
            data = self._read_file(filepath)
            logger.info(f"Read {len(data)} rows from {filepath}")
            
            # Validate
            valid, errors = DataValidator.validate_parties(data)
            logger.info(f"Validation: {len(valid)} valid, {len(errors)} errors")
            
            if not valid:
                return {
                    'status': 'failed',
                    'message': 'No valid data to import',
                    'imported': 0,
                    'errors': errors
                }
            
            # Import valid rows
            imported = []
            import_errors = []
            
            for row in valid:
                try:
                    # Check if party already exists
                    existing = self.session.query(Party).filter(Party.name == row['name']).first()
                    if existing:
                        import_errors.append(f"Party '{row['name']}' already exists (ID: {existing.id})")
                        continue
                    
                    party = Party(
                        name=row['name'],
                        phone=row.get('phone'),
                        address=row.get('address'),
                        notes=row.get('notes')
                    )
                    self.session.add(party)
                    self.session.flush()  # Get ID without committing
                    imported.append(party.id)
                    
                except Exception as e:
                    import_errors.append(f"Failed to import '{row['name']}': {str(e)}")
                    logger.error(f"Import error: {e}")
            
            # Commit all at once
            if imported:
                self.session.commit()
                logger.info(f"Imported {len(imported)} parties")
            
            # Log import
            self.import_history.add_import(
                user_id=user_id,
                entity_type='parties',
                filepath=filepath,
                imported_count=len(imported),
                error_count=len(errors) + len(import_errors),
                imported_ids=imported
            )
            
            all_errors = errors + import_errors
            
            return {
                'status': 'success' if imported else 'failed',
                'message': f'Imported {len(imported)} parties' if imported else 'No parties imported',
                'imported': len(imported),
                'errors': all_errors,
                'imported_ids': imported
            }
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Import failed: {e}")
            return {
                'status': 'error',
                'message': str(e),
                'imported': 0,
                'errors': [str(e)]
            }
    
    def import_raw_materials(self, filepath: str, user_id: int = 1) -> dict:
        """Import raw materials from file"""
        try:
            # Read file
            data = self._read_file(filepath)
            logger.info(f"Read {len(data)} rows from {filepath}")
            
            # Validate
            valid, errors = DataValidator.validate_raw_materials(data)
            logger.info(f"Validation: {len(valid)} valid, {len(errors)} errors")
            
            if not valid:
                return {
                    'status': 'failed',
                    'message': 'No valid data to import',
                    'imported': 0,
                    'errors': errors
                }
            
            # Import valid rows
            imported = []
            import_errors = []
            
            for row in valid:
                try:
                    # Check if material already exists
                    existing = self.session.query(RawMaterial).filter(RawMaterial.name == row['name']).first()
                    if existing:
                        import_errors.append(f"Raw material '{row['name']}' already exists (ID: {existing.id})")
                        continue
                    
                    # Get supplier ID if supplier name provided
                    supplier_id = None
                    if row.get('supplier_name'):
                        supplier = self.session.query(Party).filter(Party.name == row['supplier_name']).first()
                        if supplier:
                            supplier_id = supplier.id
                        else:
                            import_errors.append(f"Supplier '{row['supplier_name']}' not found for material '{row['name']}'")
                    
                    material = RawMaterial(
                        name=row['name'],
                        unit=row['unit'],
                        current_stock=0.0,
                        low_stock_alert=row['low_stock_alert'],
                        supplier_id=supplier_id,
                        notes=row.get('notes')
                    )
                    self.session.add(material)
                    self.session.flush()
                    imported.append(material.id)
                    
                except Exception as e:
                    import_errors.append(f"Failed to import '{row['name']}': {str(e)}")
                    logger.error(f"Import error: {e}")
            
            # Commit all at once
            if imported:
                self.session.commit()
                logger.info(f"Imported {len(imported)} raw materials")
            
            # Log import
            self.import_history.add_import(
                user_id=user_id,
                entity_type='raw_materials',
                filepath=filepath,
                imported_count=len(imported),
                error_count=len(errors) + len(import_errors),
                imported_ids=imported
            )
            
            all_errors = errors + import_errors
            
            return {
                'status': 'success' if imported else 'failed',
                'message': f'Imported {len(imported)} raw materials' if imported else 'No materials imported',
                'imported': len(imported),
                'errors': all_errors,
                'imported_ids': imported
            }
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Import failed: {e}")
            return {
                'status': 'error',
                'message': str(e),
                'imported': 0,
                'errors': [str(e)]
            }
    
    def import_expenses(self, filepath: str, user_id: int = 1) -> dict:
        """Import expenses from file"""
        try:
            # Read file
            data = self._read_file(filepath)
            logger.info(f"Read {len(data)} rows from {filepath}")
            
            # Validate
            valid, errors = DataValidator.validate_expenses(data)
            logger.info(f"Validation: {len(valid)} valid, {len(errors)} errors")
            
            if not valid:
                return {
                    'status': 'failed',
                    'message': 'No valid data to import',
                    'imported': 0,
                    'errors': errors
                }
            
            # Import valid rows
            imported = []
            import_errors = []
            
            for row in valid:
                try:
                    # Get farm ID if farm name provided
                    farm_id = None
                    if row.get('farm_name'):
                        farm = self.session.query(Farm).filter(Farm.name == row['farm_name']).first()
                        if farm:
                            farm_id = farm.id
                        else:
                            import_errors.append(f"Farm '{row['farm_name']}' not found for expense on {row['date']}")
                            continue
                    
                    expense = Expense(
                        date=row['date'],
                        farm_id=farm_id,
                        category=row['category'],
                        amount_afg=row['amount_afg'],
                        amount_usd=row['amount_usd'],
                        description=row.get('description'),
                        payment_method=row['payment_method']
                    )
                    self.session.add(expense)
                    self.session.flush()
                    imported.append(expense.id)
                    
                except Exception as e:
                    import_errors.append(f"Failed to import expense on {row['date']}: {str(e)}")
                    logger.error(f"Import error: {e}")
            
            # Commit all at once
            if imported:
                self.session.commit()
                logger.info(f"Imported {len(imported)} expenses")
            
            # Log import
            self.import_history.add_import(
                user_id=user_id,
                entity_type='expenses',
                filepath=filepath,
                imported_count=len(imported),
                error_count=len(errors) + len(import_errors),
                imported_ids=imported
            )
            
            all_errors = errors + import_errors
            
            return {
                'status': 'success' if imported else 'failed',
                'message': f'Imported {len(imported)} expenses' if imported else 'No expenses imported',
                'imported': len(imported),
                'errors': all_errors,
                'imported_ids': imported
            }
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Import failed: {e}")
            return {
                'status': 'error',
                'message': str(e),
                'imported': 0,
                'errors': [str(e)]
            }
    
    def import_employees(self, filepath: str, user_id: int = 1) -> dict:
        """Import employees from file"""
        try:
            # Read file
            data = self._read_file(filepath)
            logger.info(f"Read {len(data)} rows from {filepath}")
            
            # Validate
            valid, errors = DataValidator.validate_employees(data)
            logger.info(f"Validation: {len(valid)} valid, {len(errors)} errors")
            
            if not valid:
                return {
                    'status': 'failed',
                    'message': 'No valid data to import',
                    'imported': 0,
                    'errors': errors
                }
            
            # Import valid rows
            imported = []
            import_errors = []
            
            for row in valid:
                try:
                    # Check if employee already exists
                    existing = self.session.query(Employee).filter(Employee.full_name == row['full_name']).first()
                    if existing:
                        import_errors.append(f"Employee '{row['full_name']}' already exists (ID: {existing.id})")
                        continue
                    
                    employee = Employee(
                        full_name=row['full_name'],
                        job_title=row['job_title'],
                        hire_date=row.get('hire_date'),
                        salary_amount=row['salary_amount'],
                        salary_period=SalaryPeriod.MONTHLY if row['salary_period'] == 'Monthly' else SalaryPeriod.DAILY,
                        is_active=row['is_active']
                    )
                    self.session.add(employee)
                    self.session.flush()
                    imported.append(employee.id)
                    
                except Exception as e:
                    import_errors.append(f"Failed to import '{row['full_name']}': {str(e)}")
                    logger.error(f"Import error: {e}")
            
            # Commit all at once
            if imported:
                self.session.commit()
                logger.info(f"Imported {len(imported)} employees")
            
            # Log import
            self.import_history.add_import(
                user_id=user_id,
                entity_type='employees',
                filepath=filepath,
                imported_count=len(imported),
                error_count=len(errors) + len(import_errors),
                imported_ids=imported
            )
            
            all_errors = errors + import_errors
            
            return {
                'status': 'success' if imported else 'failed',
                'message': f'Imported {len(imported)} employees' if imported else 'No employees imported',
                'imported': len(imported),
                'errors': all_errors,
                'imported_ids': imported
            }
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Import failed: {e}")
            return {
                'status': 'error',
                'message': str(e),
                'imported': 0,
                'errors': [str(e)]
            }
    
    def close_session(self):
        """Close database session"""
        if self._owned_session and self.session:
            self.session.close()
