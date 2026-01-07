"""
Excel Export/Import utilities for Egg Farm Management System
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export data to Excel format"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self):
        """Create a new workbook"""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
    
    def add_sheet(self, sheet_name: str, data: List[List[Any]], headers: Optional[List[str]] = None):
        """
        Add a sheet with data to the workbook
        
        Args:
            sheet_name: Name of the sheet
            data: List of rows (each row is a list of values)
            headers: Optional header row
        """
        if self.workbook is None:
            self.create_workbook()
        
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Add headers if provided
        if headers:
            ws.append(headers)
            # Style header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
        
        # Add data rows
        for row_data in data:
            ws.append(row_data)
            # Add borders to data rows
            for cell in ws[ws.max_row]:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def save(self, file_path: Path):
        """
        Save workbook to file
        
        Args:
            file_path: Path where to save the Excel file
        """
        if self.workbook is None:
            raise ValueError("No workbook created. Call create_workbook() first.")
        
        self.workbook.save(file_path)
        logger.info(f"Excel file saved: {file_path}")
    
    def export_report_data(self, report_data: Dict[str, Any], report_type: str, file_path: Path):
        """
        Export report data to Excel
        
        Args:
            report_data: Report data dictionary
            report_type: Type of report (daily_production, monthly_production, etc.)
            file_path: Path to save Excel file
        """
        self.create_workbook()
        
        if report_type == "daily_production":
            self._export_daily_production(report_data)
        elif report_type == "monthly_production":
            self._export_monthly_production(report_data)
        elif report_type == "feed_usage":
            self._export_feed_usage(report_data)
        elif report_type == "party_statement":
            self._export_party_statement(report_data)
        else:
            # Generic export
            self.add_sheet("Report Data", [list(report_data.values())], list(report_data.keys()))
        
        self.save(file_path)
    
    def _export_daily_production(self, data: Dict):
        """Export daily production report with enhanced formatting"""
        ws = self.workbook.create_sheet(title="Daily Production")
        
        # Add header info
        date_str = data.get('date', '')
        if hasattr(date_str, 'strftime'):
            date_str = date_str.strftime('%Y-%m-%d')
        
        ws['A1'] = "Farm:"
        ws['B1'] = data.get('farm', 'N/A')
        ws['A2'] = "Date:"
        ws['B2'] = str(date_str)
        
        # Style header info
        info_font = Font(bold=True, size=12)
        ws['A1'].font = info_font
        ws['A2'].font = info_font
        
        # Add table headers
        headers = ["Shed", "Small", "Medium", "Large", "Broken", "Total", "Usable"]
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[4]:  # Header row is row 4
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for shed in data.get('sheds', []):
            row = [
                shed.get('name', ''),
                shed.get('small', 0),
                shed.get('medium', 0),
                shed.get('large', 0),
                shed.get('broken', 0),
                shed.get('total', 0),
                shed.get('usable', 0)
            ]
            ws.append(row)
            # Add borders
            for cell in ws[ws.max_row]:
                cell.border = border
        
        # Add totals row with bold formatting
        totals = data.get('totals', {})
        totals_row = [
            'TOTAL',
            totals.get('small', 0),
            totals.get('medium', 0),
            totals.get('large', 0),
            totals.get('broken', 0),
            totals.get('total', 0),
            totals.get('usable', 0)
        ]
        ws.append(totals_row)
        
        # Style totals row
        totals_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        totals_font = Font(bold=True, size=11)
        for cell in ws[ws.max_row]:
            cell.fill = totals_fill
            cell.font = totals_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _export_monthly_production(self, data: Dict):
        """Export monthly production report with enhanced formatting"""
        ws = self.workbook.create_sheet(title="Monthly Production")
        
        # Add header info
        ws['A1'] = "Farm:"
        ws['B1'] = data.get('farm', 'N/A')
        ws['A2'] = "Month:"
        ws['B2'] = f"{data.get('month', 'N/A')}/{data.get('year', 'N/A')}"
        
        # Style header info
        info_font = Font(bold=True, size=12)
        ws['A1'].font = info_font
        ws['A2'].font = info_font
        
        # Add table headers
        headers = ["Date", "Total Eggs", "Usable Eggs", "Small", "Medium", "Large", "Broken"]
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[4]:  # Header row is row 4
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for date_str, values in sorted(data.get('daily_summary', {}).items()):
            row = [
                str(date_str),
                values.get('total', 0),
                values.get('usable', 0),
                values.get('small', 0),
                values.get('medium', 0),
                values.get('large', 0),
                values.get('broken', 0)
            ]
            ws.append(row)
            # Add borders
            for cell in ws[ws.max_row]:
                cell.border = border
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _export_feed_usage(self, data: Dict):
        """Export feed usage report with enhanced formatting"""
        ws = self.workbook.create_sheet(title="Feed Usage")
        
        # Add header info
        ws['A1'] = "Farm:"
        ws['B1'] = data.get('farm', 'N/A')
        ws['A2'] = "Period:"
        ws['B2'] = f"{data.get('start_date', '')} to {data.get('end_date', '')}"
        
        # Style header info
        info_font = Font(bold=True, size=12)
        ws['A1'].font = info_font
        ws['A2'].font = info_font
        
        # Add table headers
        headers = ["Shed", "Feed Type", "Total (kg)", "Issues", "Avg per Issue (kg)", "Cost (AFG)", "Cost (USD)"]
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[4]:  # Header row is row 4
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for shed_name, shed_data in data.get('sheds', {}).items():
            row = [
                shed_name,
                shed_data.get('feed_type', ''),
                shed_data.get('total_kg', 0),
                shed_data.get('issue_count', 0),
                shed_data.get('avg_per_issue', 0),
                shed_data.get('total_cost_afg', 0),
                shed_data.get('total_cost_usd', 0)
            ]
            ws.append(row)
            # Add borders
            for cell in ws[ws.max_row]:
                cell.border = border
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _export_party_statement(self, data: Dict):
        """Export party statement with enhanced formatting"""
        ws = self.workbook.create_sheet(title="Party Statement")
        
        # Add header info
        ws['A1'] = "Party:"
        ws['B1'] = data.get('party', 'N/A')
        ws['A2'] = "Final Balance (AFG):"
        ws['B2'] = data.get('final_balance_afg', 0)
        ws['A3'] = "Final Balance (USD):"
        ws['B3'] = data.get('final_balance_usd', 0)
        
        # Style header info
        info_font = Font(bold=True, size=12)
        ws['A1'].font = info_font
        ws['A2'].font = info_font
        ws['A3'].font = info_font
        
        # Add table headers
        headers = ["Date", "Description", "Debit (AFG)", "Credit (AFG)", "Balance (AFG)", 
                   "Debit (USD)", "Credit (USD)", "Balance (USD)"]
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[5]:  # Header row is row 5
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for entry in data.get('entries', []):
            row = [
                str(entry.get('date', '')),
                entry.get('description', ''),
                entry.get('debit_afg', 0),
                entry.get('credit_afg', 0),
                entry.get('balance_afg', 0),
                entry.get('debit_usd', 0),
                entry.get('credit_usd', 0),
                entry.get('balance_usd', 0)
            ]
            ws.append(row)
            # Add borders
            for cell in ws[ws.max_row]:
                cell.border = border
        
        # Add summary row with bold formatting
        summary_row = [
            '', 'FINAL BALANCE', '', '', data.get('final_balance_afg', 0),
            '', '', data.get('final_balance_usd', 0)
        ]
        ws.append(summary_row)
        
        # Style summary row
        summary_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        summary_font = Font(bold=True, size=11)
        for cell in ws[ws.max_row]:
            cell.fill = summary_fill
            cell.font = summary_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def export_table_data(self, headers: List[str], rows: List[List[Any]], file_path: Path, sheet_name: str = "Data"):
        """
        Export table data to Excel
        
        Args:
            headers: Column headers
            rows: Data rows
            file_path: Path to save Excel file
            sheet_name: Name of the sheet
        """
        self.create_workbook()
        self.add_sheet(sheet_name, rows, headers)
        self.save(file_path)


class ExcelImporter:
    """Import data from Excel format"""
    
    @staticmethod
    def read_sheet(file_path: Path, sheet_name: Optional[str] = None) -> List[List[Any]]:
        """
        Read data from an Excel sheet
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read (None for first sheet)
            
        Returns:
            List of rows (each row is a list of values)
        """
        from openpyxl import load_workbook
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            data = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    data.append(list(row))
            
            return data
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise
    
    @staticmethod
    def read_as_dict(file_path: Path, sheet_name: Optional[str] = None, header_row: int = 0) -> List[Dict[str, Any]]:
        """
        Read Excel sheet as list of dictionaries (first row as keys)
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read
            header_row: Row index containing headers (0-based)
            
        Returns:
            List of dictionaries
        """
        from openpyxl import load_workbook
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Read headers
            headers = []
            for cell in ws[header_row + 1]:
                headers.append(str(cell.value) if cell.value else f"Column{len(headers) + 1}")
            
            # Read data
            data = []
            for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = value
                    data.append(row_dict)
            
            return data
        except Exception as e:
            logger.error(f"Failed to read Excel file as dict: {e}")
            raise

